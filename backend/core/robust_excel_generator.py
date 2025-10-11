import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import json

class RobustExcelGenerator:
    def __init__(self):
        self.workbook = openpyxl.Workbook()
    
    def safe_get_value(self, data, key, default=""):
        """Safely get value from data, handling None and missing keys"""
        try:
            value = data.get(key, default)
            if value is None:
                return default
            return str(value)
        except:
            return default
    
    def safe_json_dumps(self, data):
        """Safely convert data to JSON string"""
        try:
            if data is None:
                return "N/A"
            return json.dumps(data, indent=2, default=str)
        except:
            return str(data) if data else "N/A"
    
    def create_summary_sheet(self, validation_results):
        """Create a simple summary sheet"""
        try:
            ws = self.workbook.active
            ws.title = "Summary"
            
            # Header
            ws['A1'] = "Data Validation Summary Report"
            ws['A1'].font = Font(bold=True, size=16)
            
            # Timestamp
            ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True)
            
            # Headers
            headers = ['Table Name', 'Source Count', 'Target Count', 'Matched', 'Changed', 'Missing', 'Extra', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Data
            row = 5
            for result in validation_results:
                ws.cell(row=row, column=1, value=self.safe_get_value(result, 'table'))
                ws.cell(row=row, column=2, value=self.safe_get_value(result, 'source_count', 0))
                ws.cell(row=row, column=3, value=self.safe_get_value(result, 'target_count', 0))
                ws.cell(row=row, column=4, value=self.safe_get_value(result, 'matched_count', 0))
                ws.cell(row=row, column=5, value=self.safe_get_value(result, 'changed_count', 0))
                ws.cell(row=row, column=6, value=self.safe_get_value(result, 'missing_in_target', 0))
                ws.cell(row=row, column=7, value=self.safe_get_value(result, 'extra_in_target', 0))
                ws.cell(row=row, column=8, value=self.safe_get_value(result, 'status'))
                row += 1
            
            # Set column widths manually
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 15
            
        except Exception as e:
            print(f"Error creating summary sheet: {e}")
            # Create a basic sheet if there's an error
            ws = self.workbook.active
            ws.title = "Summary"
            ws['A1'] = "Error generating report"
            ws['A2'] = str(e)
    
    def create_detailed_sheet(self, detailed_records, table_name):
        """Create detailed differences sheet"""
        try:
            ws = self.workbook.create_sheet("Detailed Differences")
            
            # Header
            ws['A1'] = f"Detailed Differences - {table_name}"
            ws['A1'].font = Font(bold=True, size=16)
            
            # Headers
            headers = ['Record ID', 'Name', 'Email', 'Difference Type', 'Source Data', 'Target Data', 'Differences', 'Sync Status', 'Timestamp']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Data
            row = 4
            for record in detailed_records:
                try:
                    # Safely extract data
                    record_id = self.safe_get_value(record, 'record_id')
                    
                    # Get name from source or target data
                    source_data = record.get('source_data', {}) or {}
                    target_data = record.get('target_data', {}) or {}
                    name = self.safe_get_value(source_data, 'name') or self.safe_get_value(target_data, 'name')
                    email = self.safe_get_value(source_data, 'email') or self.safe_get_value(target_data, 'email')
                    
                    difference_type = self.safe_get_value(record, 'difference_type')
                    differences = record.get('differences', [])
                    if isinstance(differences, list):
                        differences_str = ', '.join([str(d) for d in differences])
                    else:
                        differences_str = str(differences)
                    
                    sync_status = 'Synced' if difference_type != 'matched' else 'No Action Needed'
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Write to Excel
                    ws.cell(row=row, column=1, value=record_id)
                    ws.cell(row=row, column=2, value=name)
                    ws.cell(row=row, column=3, value=email)
                    ws.cell(row=row, column=4, value=difference_type)
                    ws.cell(row=row, column=5, value=self.safe_json_dumps(source_data))
                    ws.cell(row=row, column=6, value=self.safe_json_dumps(target_data))
                    ws.cell(row=row, column=7, value=differences_str)
                    ws.cell(row=row, column=8, value=sync_status)
                    ws.cell(row=row, column=9, value=timestamp)
                    
                    row += 1
                    
                except Exception as e:
                    print(f"Error processing record {row}: {e}")
                    # Add error row
                    ws.cell(row=row, column=1, value=f"Error in record {row}")
                    ws.cell(row=row, column=2, value=str(e))
                    row += 1
            
            # Set column widths manually
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 40
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 20
            
        except Exception as e:
            print(f"Error creating detailed sheet: {e}")
            # Create error sheet
            ws = self.workbook.create_sheet("Error")
            ws['A1'] = "Error creating detailed differences sheet"
            ws['A2'] = str(e)
    
    def generate_excel_report(self, validation_results, detailed_records, table_name):
        """Generate Excel report with summary and detailed sheets"""
        try:
            self.create_summary_sheet(validation_results)
            self.create_detailed_sheet(detailed_records, table_name)
            return self.workbook
        except Exception as e:
            print(f"Error generating Excel report: {e}")
            # Return basic workbook with error
            ws = self.workbook.active
            ws.title = "Error"
            ws['A1'] = "Error generating Excel report"
            ws['A2'] = str(e)
            return self.workbook
    
    def save_report(self, filename):
        """Save the Excel report to file"""
        try:
            self.workbook.save(filename)
            return filename
        except Exception as e:
            print(f"Error saving Excel report: {e}")
            raise e
